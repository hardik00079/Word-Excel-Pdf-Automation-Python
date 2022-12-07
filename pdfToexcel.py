import PyPDF2
import os
import re
import openpyxl

your_excel = openpyxl.load_workbook('Output_from_pdf.xlsx')
your_sheet = your_excel['Sheet1']

for file_name in  os.listdir('PDF'):
    print(file_name)
    
    load_pdf = open(r'E:\Python Workspace\PdfToExcelPythonProject\\PDF\\'+file_name,'rb')
    read_pdf=PyPDF2.PdfFileReader(load_pdf)
    page_count = read_pdf.getNumPages()
    first_page = read_pdf.getPage(0)
    page_content = first_page.extract_text()
    
    mobile_number = re.search(r'(?:\+?\d{2}[ -]?)?\d{9}',page_content).group()
    print(mobile_number)    
    email_id= re.search(r'([A-Za-z0-9._-]+@[A-Za-z0-9._-]+\.[A-Za-z0-9._-]+)',page_content).group()
    print(email_id)
    
    reg_1 = '(?<=Dear, )(.*)'
    found_name = re.search("|".join([reg_1]),page_content).group()
    print(found_name)
    reg_2 = '(?<=• Address – )(.*)'
    found_address = re.search("|".join([reg_2]),page_content).group()
    print(found_address)
    reg_3 = '(?<=• Mobile Number – )(.*)'
    found_monumber = re.search("|".join([reg_3]),page_content).group()
    print(found_monumber)
    
    reg_4 = '(?<=• College – )(.*)'
    college = re.search("|".join([reg_4]),page_content).group()
    print(college)
    
    last_row_number= your_sheet.max_row
    print(last_row_number)
    
    your_sheet.cell(column=1,row=last_row_number+1).value = found_name
    your_sheet.cell(column=2,row=last_row_number+1).value = found_monumber
    your_sheet.cell(column=3,row=last_row_number+1).value = email_id
    your_sheet.cell(column=4,row=last_row_number+1).value = found_address
    
    your_excel.save('Output_from_pdf.xlsx')